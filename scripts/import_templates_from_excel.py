#!/usr/bin/env python3
"""Import GA4 audit templates from the mapping workbook into Neon/Postgres.

Default mode is a dry run. Set NEON_DATABASE_URL and pass --execute to write.
"""

from __future__ import annotations

import argparse
import hashlib
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import openpyxl


SHEET_CONFIG = {
    "Jagran Hindi": ("www.jagran.com", "G-3RLQSM7QQQ", "GTM-5CTQK3", ""),
    "Thedailyjagran.com": ("www.thedailyjagran.com", "G-171L9K03BG", "GTM-5CTQK3", ""),
    "onlymyhealth.com English": ("www.onlymyhealth.com", "G-7REYEWS85G", "GTM-5LTRVCK", "OnlyMyHealth English"),
    "onlymyhealth.com Hindi": ("www.onlymyhealth.com", "G-7REYEWS85G", "GTM-5LTRVCK", "OnlyMyHealth Hindi"),
    "Herzindagi Hindi": ("www.herzindagi.com", "G-WHCRWZ50HH", "GTM-WWVXM33", "Herzindagi Hindi"),
    "Herzindagi English": ("www.herzindagi.com", "G-WHCRWZ50HH", "GTM-WWVXM33", "Herzindagi English"),
    "jagranjosh.com": ("www.jagranjosh.com", "G-8J9SC9WB3T", "GTM-N62LNQ", ""),
    "punjabijagran.com": ("www.punjabijagran.com", "G-18RV68H7Y1", "GTM-5CTQK3", ""),
    "new sheet events": ("www.jagran.com", "G-3RLQSM7QQQ", "GTM-5CTQK3", ""),
}

FIELD_ALIASES = {
    "article_type": "article_type",
    "author": "author",
    "category": "category",
    "dynamic_video_embed_type": "dynamic_video_embed_type",
    "embed_type": "embed_type",
    "event": "event name",
    "event name": "event name",
    "event_name": "event name",
    "language": "Language",
    "online_offline": "online_offline",
    "page_type": "page_type",
    "player_type": "player_type",
    "planned_trending": "planned_trending",
    "position_fold": "position_fold",
    "posted_by": "posted_by",
    "publish_date": "publish_date",
    "registration_status": "registration_status",
    "scroll_percent": "scroll_percent",
    "section_name": "section_name",
    "smart_view_enabled": "smart_view_enabled",
    "story_id": "story_id",
    "sub_category": "sub_category",
    "sub_sub_category": "sub_sub_category",
    "tags": "tags",
    "tvc_event_name": "tvc_event_name",
    "update_date": "update_date",
    "user id": "User_Id",
    "user_id": "User_Id",
    "user_status": "User_Status",
    "usertype": "usertype",
    "video_duration": "Video_duration",
    "video_orientation": "video_orientation",
    "video_percent": "video_percent",
    "video_title": "Video_title",
    "word_count": "word_count",
}

DATETIME_RE = r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}(?:[+-]\d{2}:\d{2}|Z)$"
INTEGER_RE = r"^\d+$"


def clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def domain_key(value: str) -> str:
    text = clean(value).lower()
    text = re.sub(r"^https?://", "", text).split("/", 1)[0]
    return text.removeprefix("www.")


def canonical_field(value) -> str:
    text = clean(value)
    key = text.lower().replace(" ", "_")
    return FIELD_ALIASES.get(text.lower()) or FIELD_ALIASES.get(key, "")


def dynamic_placeholder(value: str) -> bool:
    text = clean(value)
    if not text:
        return False
    tokens = [
        re.sub(r"[{}\s]+", "", item.lower())
        for item in re.split(r"[,|/]+", text)
        if re.sub(r"[{}\s]+", "", item.lower())
    ]
    return bool(tokens) and all(item in {"dynamic", "dynamci"} for item in tokens)


def normalize_expected(value: str) -> str:
    text = clean(value)
    text = re.sub(r"^\(?\s*", "", text)
    text = re.sub(r"\s*\)?$", "", text)
    text = re.sub(r"^static\s*[-:]\s*", "", text, flags=re.I)
    if dynamic_placeholder(text):
        return "dynamic"
    return {
        "articl detail": "article detail",
        "dynamci": "dynamic",
        "video_interation": "video_interaction",
    }.get(text.lower(), clean(text))


def looks_event_label(value: str) -> bool:
    text = clean(value).lower()
    if not text or text in {"detail", "listing", "landing", "lisitng/landing", "listing/landing", "landing / listing"}:
        return False
    return text in {"video_interaction", "viddeo_interaction"} or text.startswith("page_") or text.endswith("_pv")


def split_events(value: str) -> list[str]:
    events = []
    for token in re.split(r"[,|\n]", clean(value)):
        event = normalize_expected(token)
        if event and event.lower() not in {"dynamic", "static", "not available", "na"}:
            events.append("video_interaction" if event == "viddeo_interaction" else event)
    return events


def allowed_values(structure: str, sample: str) -> list[str]:
    structure_text = clean(structure)
    lower = structure_text.lower()
    if dynamic_placeholder(structure_text):
        return []
    values = []
    if "yes/no" in lower:
        values.extend(["yes", "no"])
    if "logged_in" in lower:
        values.append("logged_in")
    if "guest" in lower:
        values.append("guest")
    match = re.search(r"\(([^)]+)\)", structure_text)
    if match:
        for token in re.split(r"[/,|&-]", match.group(1)):
            item = normalize_expected(token)
            if item and item.lower() not in {"dynamic", "static"}:
                values.append(item)
    if ("dynamic" in lower or "dynamci" in lower) and values:
        item = normalize_expected(sample)
        if item and item.lower() not in {"dynamic", "static", "na", "not available"}:
            values.append(item)
    if not values and "/" in structure_text:
        for token in structure_text.split("/"):
            item = normalize_expected(token)
            if item and item.lower() not in {"dynamic", "static", "na", "not available"}:
                values.append(item)
    seen = set()
    unique = []
    for value in values:
        key = value.lower()
        if key not in seen:
            seen.add(key)
            unique.append(value)
    return unique


def infer_rule(field: str, sample: str, structure: str, page_type: str) -> dict:
    field = clean(field)
    sample_value = normalize_expected(sample)
    structure_text = clean(structure)
    structure_value = normalize_expected(structure_text)
    lower = structure_value.lower()

    if field == "event name":
        event = sample_value or structure_text or "page_view"
        return {"rule_scope": "event", "field_name": event, "rule_type": "exact", "expected_values": event, "notes": "Imported from GA mapping Excel."}
    if field == "page_type":
        expected = clean(page_type) or sample_value or structure_text
        return {"rule_scope": "execution", "field_name": "page_type", "rule_type": "exact", "expected_values": expected, "notes": f"Imported from GA mapping Excel. Sample: {sample_value or '-'}"}
    if field in {"publish_date", "update_date"} or "yyyy-mm-dd" in lower:
        return {"rule_scope": "execution", "field_name": field, "rule_type": "regex", "expected_values": DATETIME_RE, "notes": f"Imported from GA mapping Excel. Expected format: {structure_text or 'ISO datetime'}"}
    if dynamic_placeholder(sample) or dynamic_placeholder(structure):
        return {"rule_scope": "execution", "field_name": field, "rule_type": "required", "expected_values": "", "notes": f"Imported from GA mapping Excel. Dynamic value. Sample: {sample_value or '-'}"}
    if field in {"story_id", "word_count"}:
        return {"rule_scope": "execution", "field_name": field, "rule_type": "regex", "expected_values": INTEGER_RE, "notes": f"Imported from GA mapping Excel. Numeric dynamic value. Sample: {sample_value or '-'}"}
    choices = allowed_values(structure_text, sample_value)
    if choices:
        return {"rule_scope": "execution", "field_name": field, "rule_type": "one_of", "expected_values": "|".join(choices), "notes": f"Imported from GA mapping Excel. Allowed by format: {structure_text}"}
    if lower == "static":
        return {"rule_scope": "execution", "field_name": field, "rule_type": "exact", "expected_values": sample_value, "notes": "Imported from GA mapping Excel. Static value."}
    if lower == "dynamic" or lower.startswith("dynamic"):
        return {"rule_scope": "execution", "field_name": field, "rule_type": "required", "expected_values": "", "notes": f"Imported from GA mapping Excel. Dynamic value. Sample: {sample_value or '-'}"}
    if structure_text:
        return {"rule_scope": "execution", "field_name": field, "rule_type": "exact", "expected_values": structure_value or structure_text, "notes": f"Imported from GA mapping Excel. Sample: {sample_value or '-'}"}
    return {"rule_scope": "execution", "field_name": field, "rule_type": "required", "expected_values": "", "notes": f"Imported from GA mapping Excel. Sample: {sample_value or '-'}"}


def merge_rule(existing: dict | None, incoming: dict, page_type: str) -> dict:
    if not existing:
        return incoming
    if clean(existing.get("field_name") or incoming.get("field_name")) == "page_type":
        existing["rule_type"] = "exact"
        existing["expected_values"] = clean(page_type)
        return existing
    if existing["rule_type"] == incoming["rule_type"] and existing.get("expected_values", "") == incoming.get("expected_values", ""):
        return existing
    if existing["rule_type"] in {"required", "optional", "regex"} or incoming["rule_type"] in {"required", "optional", "regex"}:
        existing["rule_type"] = "required"
        existing["expected_values"] = ""
        existing["notes"] = "Imported from GA mapping Excel. Multiple examples differ, so any non-empty value is accepted."
        return existing
    values = []
    for item in [*str(existing.get("expected_values") or "").split("|"), *str(incoming.get("expected_values") or "").split("|")]:
        value = normalize_expected(item)
        if value and value.lower() not in {entry.lower() for entry in values}:
            values.append(value)
    existing["rule_type"] = "one_of" if len(values) > 1 else "exact"
    existing["expected_values"] = "|".join(values)
    existing["notes"] = "Imported from GA mapping Excel. Multiple static examples were merged."
    return existing


def infer_url_patterns(page_type: str, urls: list[str], domain: str) -> str:
    page_type_lower = clean(page_type).lower()
    patterns = []
    if domain_key(domain) == "jagran.com":
        if page_type_lower == "article detail":
            patterns.extend(["https://www.jagran.com/*/*-*.html", "https://www.jagran.com/smart-choice/*/*"])
        elif page_type_lower == "live blog detail":
            patterns.append("https://www.jagran.com/*/*-lb-*.html")
        elif page_type_lower == "short video detail":
            patterns.append("https://www.jagran.com/short-videos/*")
        elif page_type_lower == "photo detail":
            patterns.append("https://www.jagran.com/photo-stories/*")
        elif page_type_lower == "topic listing page":
            patterns.append("https://www.jagran.com/topics/*")
    for url in urls:
        item = clean(url)
        if item and item not in patterns:
            patterns.append(item)
    return "\n".join(patterns)


def template_id_for(domain: str, template_name: str) -> str:
    digest = hashlib.sha1(f"{domain_key(domain)}|{template_name.lower()}".encode()).hexdigest()[:12]
    return f"tpl_{digest}"


def rule_id_for(template_id: str, index: int) -> str:
    return f"rule_{template_id.removeprefix('tpl_')}_{index:03d}"


def page_type_from_row(row: list) -> str:
    for idx in range(3, len(row)):
        if canonical_field(row[idx]) == "page_type":
            return clean(row[idx + 1] if idx + 1 < len(row) else "") or clean(row[idx + 2] if idx + 2 < len(row) else "")
    return ""


def parse_sheet(ws, config: tuple[str, str, str, str]) -> list[dict]:
    domain, measurement_id, container_id, prefix = config
    templates = {}
    current_section = ""
    start_row = 2 if clean(ws.cell(1, 1).value).lower() == "sections" else 3
    for row_values in ws.iter_rows(min_row=start_row, values_only=True):
        row = list(row_values)
        section = clean(row[0] if row else "")
        if section and not looks_event_label(section):
            current_section = section
        page_type = clean(row[1] if len(row) > 1 else "") or page_type_from_row(row)
        if not page_type:
            continue
        template_name = f"{prefix} - {page_type}" if prefix else page_type
        key = (domain_key(domain), template_name.lower())
        template = templates.setdefault(
            key,
            {
                "template_name": template_name,
                "domain_name": domain,
                "measurement_id": measurement_id,
                "container_id": container_id,
                "url_examples": [],
                "rules": {},
            },
        )
        for url in re.split(r"[\r\n]+", str(row[2] if len(row) > 2 and row[2] is not None else "")):
            url = clean(url)
            if url and url not in template["url_examples"]:
                template["url_examples"].append(url)
        page_view = {"rule_scope": "event", "field_name": "page_view", "rule_type": "exact", "expected_values": "page_view", "notes": "Added by importer because every GA4 audit template should validate page_view."}
        template["rules"][("event", "page_view")] = merge_rule(template["rules"].get(("event", "page_view")), page_view, page_type)
        if looks_event_label(section or current_section):
            for event in split_events(section or current_section):
                rule = {"rule_scope": "event", "field_name": event, "rule_type": "exact", "expected_values": event, "notes": "Imported from GA mapping Excel."}
                template["rules"][("event", event.lower())] = merge_rule(template["rules"].get(("event", event.lower())), rule, page_type)
        seen = set()
        col = 3
        while col < len(row):
            field = canonical_field(row[col])
            if not field:
                col += 1
                continue
            nxt = col + 1
            sample = ""
            structure = ""
            if nxt < len(row) and not canonical_field(row[nxt]):
                sample = clean(row[nxt])
                nxt += 1
            if nxt < len(row) and not canonical_field(row[nxt]):
                structure = clean(row[nxt])
                nxt += 1
            if field == "page_type" and field in seen and sample.lower() != page_type.lower():
                break
            seen.add(field)
            rule = infer_rule(field, sample, structure, page_type)
            rule_key = (rule["rule_scope"].lower(), rule["field_name"].lower())
            template["rules"][rule_key] = merge_rule(template["rules"].get(rule_key), rule, page_type)
            if field == "tvc_event_name":
                for event in split_events(sample or structure):
                    event_rule = {"rule_scope": "event", "field_name": event, "rule_type": "exact", "expected_values": event, "notes": "Imported from GA mapping Excel."}
                    template["rules"][("event", event.lower())] = merge_rule(template["rules"].get(("event", event.lower())), event_rule, page_type)
            col = max(nxt, col + 1)
    output = []
    for template in templates.values():
        output.append(
            {
                "template_name": template["template_name"],
                "domain_name": template["domain_name"],
                "measurement_id": template["measurement_id"],
                "container_id": template["container_id"],
                "url_pattern": infer_url_patterns(template["template_name"], template["url_examples"], template["domain_name"]),
                "rules": list(template["rules"].values()),
            }
        )
    return sorted(output, key=lambda item: (domain_key(item["domain_name"]), item["template_name"].lower()))


def parse_workbook(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    templates = []
    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name in wb.sheetnames:
            templates.extend(parse_sheet(wb[sheet_name], config))
    return merge_generated_templates(templates)


def merge_generated_templates(templates: list[dict]) -> list[dict]:
    merged = {}
    for template in templates:
        key = (domain_key(template["domain_name"]), template["template_name"].lower())
        existing = merged.get(key)
        if not existing:
            merged[key] = {
                **template,
                "rules": list(template.get("rules") or []),
            }
            continue
        existing_urls = [clean(line) for line in str(existing.get("url_pattern") or "").splitlines() if clean(line)]
        for line in str(template.get("url_pattern") or "").splitlines():
            line = clean(line)
            if line and line not in existing_urls:
                existing_urls.append(line)
        existing["url_pattern"] = "\n".join(existing_urls)

        rules_by_key = {
            (rule.get("rule_scope", "").lower(), rule.get("field_name", "").lower()): rule
            for rule in existing.get("rules") or []
        }
        for rule in template.get("rules") or []:
            rule_key = (rule.get("rule_scope", "").lower(), rule.get("field_name", "").lower())
            rules_by_key[rule_key] = merge_rule(rules_by_key.get(rule_key), rule, template["template_name"])
        existing["rules"] = list(rules_by_key.values())
    return sorted(merged.values(), key=lambda item: (domain_key(item["domain_name"]), item["template_name"].lower()))


def connect(database_url: str):
    import psycopg
    from psycopg.rows import dict_row

    return psycopg.connect(database_url, row_factory=dict_row)


def load_existing(conn):
    with conn.cursor() as cur:
        cur.execute("SELECT * FROM ga_audit_templates")
        templates = cur.fetchall()
        cur.execute("SELECT * FROM ga_audit_template_rules")
        rules = cur.fetchall()
    by_key = {}
    for template in templates:
        by_key[(domain_key(template.get("domain_name")), clean(template.get("template_name")).lower())] = template
    return templates, rules, by_key


def import_templates(conn, generated: list[dict], actor: str):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    existing_templates, existing_rules, existing_by_key = load_existing(conn)
    generated_keys = {(domain_key(t["domain_name"]), t["template_name"].lower()) for t in generated}
    existing_rule_ids = defaultdict(list)
    for rule in existing_rules:
        existing_rule_ids[rule.get("template_id")].append(rule.get("rule_id"))

    created = updated = deleted_rules = inserted_rules = 0
    with conn.cursor() as cur:
        for template in generated:
            key = (domain_key(template["domain_name"]), template["template_name"].lower())
            existing = existing_by_key.get(key)
            template_id = existing.get("template_id") if existing else template_id_for(template["domain_name"], template["template_name"])
            if existing:
                cur.execute(
                    """
                    UPDATE ga_audit_templates
                    SET domain_name=%s, measurement_id=%s, container_id=%s, url_pattern=%s, active=true
                    WHERE template_id=%s
                    """,
                    (template["domain_name"], template["measurement_id"], template["container_id"], template["url_pattern"], template_id),
                )
                updated += 1
            else:
                cur.execute(
                    """
                    INSERT INTO ga_audit_templates (
                        template_id, template_name, domain_name, measurement_id, container_id,
                        url_pattern, active, created_by, created_at
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,true,%s,%s)
                    """,
                    (template_id, template["template_name"], template["domain_name"], template["measurement_id"], template["container_id"], template["url_pattern"], actor, now),
                )
                created += 1
            if existing_rule_ids.get(template_id):
                cur.execute("DELETE FROM ga_audit_template_rules WHERE template_id=%s", (template_id,))
                deleted_rules += len(existing_rule_ids[template_id])
            for index, rule in enumerate(template["rules"], start=1):
                cur.execute(
                    """
                    INSERT INTO ga_audit_template_rules (
                        rule_id, template_id, rule_scope, field_name, rule_type,
                        expected_values, notes, created_by, created_at
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        rule_id_for(template_id, index),
                        template_id,
                        rule["rule_scope"],
                        rule["field_name"],
                        rule["rule_type"],
                        rule.get("expected_values", ""),
                        rule.get("notes", ""),
                        actor,
                        now,
                    ),
                )
                inserted_rules += 1
        conn.commit()
    return {
        "created_templates": created,
        "updated_templates": updated,
        "deleted_rules": deleted_rules,
        "inserted_rules": inserted_rules,
        "generated_templates": len(generated),
        "generated_rules": sum(len(t["rules"]) for t in generated),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--execute", action="store_true")
    parser.add_argument("--actor", default="excel-import")
    args = parser.parse_args()

    generated = parse_workbook(args.workbook)
    by_domain = Counter(domain_key(item["domain_name"]) for item in generated)
    rule_by_domain = defaultdict(int)
    for item in generated:
        rule_by_domain[domain_key(item["domain_name"])] += len(item["rules"])
    print("Generated templates:")
    for key in sorted(by_domain):
        print(f"  {key}: {by_domain[key]} template(s), {rule_by_domain[key]} rule(s)")
    print(f"Total: {len(generated)} template(s), {sum(rule_by_domain.values())} rule(s)")

    if not args.execute:
        print("Dry run only. Set NEON_DATABASE_URL and pass --execute to upload.")
        return 0
    database_url = os.environ.get("NEON_DATABASE_URL") or os.environ.get("DATABASE_URL") or os.environ.get("POSTGRES_URL")
    if not database_url:
        print("NEON_DATABASE_URL/DATABASE_URL/POSTGRES_URL is required for --execute.", file=sys.stderr)
        return 2
    with connect(database_url) as conn:
        summary = import_templates(conn, generated, args.actor)
    print("Upload complete:")
    for key, value in summary.items():
        print(f"  {key}: {value}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
